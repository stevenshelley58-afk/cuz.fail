from __future__ import annotations

from io import BytesIO

from docx import Document
from draftcheck_core.json_utils import from_json, to_json
from draftcheck_core.models import Export, HumanSignoff, utcnow
from draftcheck_core.object_storage import StoredObject
from draftcheck_core.project_service import _apply_export_signoff
from openpyxl import load_workbook

from tests.test_sources_retrieval import seed_source


def create_project(client):
    response = client.post(
        "/v1/projects",
        json={
            "project_name": "Example RFI job",
            "address": "1 Example Street, Spearwood WA",
            "local_government": "Cockburn",
            "lot_plan": "Lot 1 on P12345",
            "project_type": "single_house_addition",
            "stage": "RFI",
            "r_code_density": "R30",
            "ncc_edition": "NCC 2022",
            "created_by": "tester",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def add_measurements(client, project_id):
    for payload in [
        {"key": "site_area_m2", "value": 500, "unit": "m2"},
        {"key": "building_footprint_m2", "value": 250, "unit": "m2"},
        {"key": "open_space_m2", "value": 200, "unit": "m2"},
        {"key": "front_setback_m", "value": 3.5, "unit": "m"},
        {"key": "garage_width_m", "value": 6, "unit": "m"},
        {"key": "frontage_width_m", "value": 12, "unit": "m"},
    ]:
        response = client.post(f"/v1/projects/{project_id}/measurements", json=payload)
        assert response.status_code == 200, response.text


def seed_building_height_source(client):
    response = client.post(
        "/v1/sources/seed",
        json={
            "title": "Accepted Project Chat R-Code Building Height Fixture",
            "jurisdiction": "WA",
            "authority": "Department of Planning, Lands and Heritage",
            "source_type": "r_code",
            "canonical_url": "https://example.test/project-chat-r-code-building-height",
            "licence_notes": "Public source fixture.",
            "access_type": "public",
            "review_status": "accepted",
            "content": "\n".join(
                [
                    "C3.2.1 Building height",
                    "For R40 coded lots, a single house may be two storeys where other standards are satisfied.",
                ]
            ),
            "version_label": "accepted",
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_compliance_matrix_uses_deterministic_calculations_and_citations(client):
    seed_source(client)
    project = create_project(client)
    add_measurements(client, project["id"])

    response = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert response.status_code == 200, response.text
    matrix = response.json()
    assert matrix["liability_notice"].startswith("Assistive drafting review only")
    assert matrix["as_of_date"]
    assert matrix["assessment_basis"] == "current_rules"
    statuses = {row["check_key"]: row["status"] for row in matrix["results"]}
    assert statuses["site_cover"] == "unsupported"
    assert statuses["open_space"] == "unsupported"
    assert statuses["front_setback"] == "unsupported"
    assert statuses["garage_dominance"] == "unsupported"
    site_cover = next(row for row in matrix["results"] if row["check_key"] == "site_cover")
    assert site_cover["as_of_date"] == matrix["as_of_date"]
    assert site_cover["assessment_basis"] == "current_rules"
    assert site_cover["proposed"].startswith("50.0%")
    assert "approved rule row for site_cover" in site_cover["missing_information"]
    assert "approved source citation" in site_cover["missing_information"]
    assert site_cover["citations"] == []
    assert site_cover["decision_trace_id"]
    assert site_cover["rule_ids"] == []
    assert site_cover["resolved_rule_ids"] == []
    assert site_cover["measurement_ids"]
    trace = client.get(f"/v1/projects/{project['id']}/checks/{site_cover['id']}/decision-trace")
    assert trace.status_code == 200, trace.text
    trace_body = trace.json()
    assert trace_body["id"] == site_cover["decision_trace_id"]
    assert trace_body["inputs"]["check_key"] == "site_cover"
    assert trace_body["measurement_ids"]
    assert trace_body["rounding_policy"].startswith("round percentage")
    assert trace_body["precedence_trace"]["resolved_rule_support"] == "not_available"
    patch = client.patch(
        f"/v1/projects/{project['id']}/checks/{site_cover['id']}",
        json={"status": "likely_pass"},
    )
    assert patch.status_code == 400
    unsafe_status_patch = client.patch(
        f"/v1/projects/{project['id']}/checks/{site_cover['id']}",
        json={"status": "missing_info"},
    )
    assert unsafe_status_patch.status_code == 400
    assert "DecisionTrace" in unsafe_status_patch.json()["detail"]
    unsafe_proposed_patch = client.patch(
        f"/v1/projects/{project['id']}/checks/{site_cover['id']}",
        json={"proposed": "manual override"},
    )
    assert unsafe_proposed_patch.status_code == 400
    assert "DecisionTrace" in unsafe_proposed_patch.json()["detail"]
    assert all(row["requires_human_review"] for row in matrix["results"])
    cited = [row for row in matrix["results"] if row["citations"]]
    assert cited == []
    canonical_matrix = client.get(f"/v1/projects/{project['id']}/compliance/matrix")
    assert canonical_matrix.status_code == 200, canonical_matrix.text
    canonical_body = canonical_matrix.json()
    assert canonical_body["project_id"] == project["id"]
    assert canonical_body["check_run_id"] == matrix["check_run_id"]
    assert canonical_body["status"] == "completed"
    assert canonical_body["as_of_date"] == matrix["as_of_date"]
    assert canonical_body["assessment_basis"] == matrix["assessment_basis"]
    assert canonical_body["requires_human_signoff"] is True
    assert canonical_body["liability_notice"].startswith("Assistive drafting review only")
    assert canonical_body["source_version_ids"] == []
    assert {row["id"] for row in canonical_body["results"]} == {row["id"] for row in matrix["results"]}

    second_run = client.post(f"/v1/projects/{project['id']}/compliance/run")
    assert second_run.status_code == 200, second_run.text
    second_matrix = second_run.json()
    latest_matrix = client.get(f"/v1/projects/{project['id']}/compliance/matrix")
    assert latest_matrix.status_code == 200, latest_matrix.text
    latest_body = latest_matrix.json()
    assert latest_body["check_run_id"] == second_matrix["check_run_id"]
    assert {row["id"] for row in latest_body["results"]} == {row["id"] for row in second_matrix["results"]}
    assert not ({row["id"] for row in latest_body["results"]} & {row["id"] for row in matrix["results"]})

    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["compliance_matrix"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    download = client.get(f"/v1/projects/{project['id']}/exports/{export.json()['id']}/download")
    assert download.status_code == 200, download.text
    exported_rows = download.json()["sections"]["compliance_matrix"]
    assert {row["id"] for row in exported_rows} == {row["id"] for row in second_matrix["results"]}
    assert not ({row["id"] for row in exported_rows} & {row["id"] for row in matrix["results"]})


def test_exports_use_exports_storage_bucket(client, monkeypatch):
    storage = RecordingStorage("exports")
    monkeypatch.setattr("draftcheck_export.service.get_export_storage", lambda: storage)
    project = create_project(client)

    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": [], "created_by": "tester"},
    )

    assert export.status_code == 200, export.text
    body = export.json()
    assert body["object_key"].startswith("s3://exports/")
    assert not body["object_key"].startswith("s3://uploads/")
    assert body["object_key"] in storage.objects
    download = client.get(f"/v1/projects/{project['id']}/exports/{body['id']}/download")
    assert download.status_code == 200, download.text


def test_compliance_matrix_covers_requested_check_pack_and_missing_info(client):
    seed_source(client)
    project = create_project(client)

    response = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert response.status_code == 200, response.text
    rows = response.json()["results"]
    statuses = {row["check_key"]: row["status"] for row in rows}
    required_keys = {
        "front_setback",
        "side_setback",
        "rear_setback",
        "site_cover",
        "open_space",
        "deep_soil_tree_planting",
        "garage_dominance",
        "street_surveillance",
        "outdoor_living_area",
        "solar_access",
        "privacy",
        "overshadowing",
        "vehicle_access",
        "bin_storage",
        "ancillary_dwelling_trigger",
        "retaining_fill_trigger",
        "bal_bushfire_trigger",
        "heritage_overlay_trigger",
        "boundary_wall",
        "title_block_completeness",
        "revision_completeness",
        "north_point_completeness",
        "scale_completeness",
        "dimension_completeness",
    }
    assert required_keys.issubset(statuses.keys())
    assert statuses["side_setback"] == "missing_info"
    assert statuses["bal_bushfire_trigger"] == "missing_info"


def test_project_chat_summarises_latest_compliance_run_for_design_question(client):
    seed_source(client)
    project = create_project(client)
    add_measurements(client, project["id"])

    run = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert run.status_code == 200, run.text

    answer = client.post(
        f"/v1/projects/{project['id']}/chat",
        json={"message": "Does my design comply?"},
    )

    assert answer.status_code == 200, answer.text
    body = answer.json()
    assert body["status"] in {"missing_info", "needs_human_review", "unsupported"}
    assert body["citations"] == []
    assert "latest deterministic compliance run" in body["answer"]
    assert "unsupported" in body["answer"]
    assert "approved rule row" in " ".join(body["missing_information"])


def test_project_chat_answers_targeted_check_from_latest_run(client):
    seed_source(client)
    project = create_project(client)
    add_measurements(client, project["id"])

    run = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert run.status_code == 200, run.text

    answer = client.post(
        f"/v1/projects/{project['id']}/chat",
        json={"message": "What did the site cover check say?"},
    )

    assert answer.status_code == 200, answer.text
    body = answer.json()
    assert body["status"] == "unsupported"
    assert "Site cover" in body["answer"]
    assert "50.0%" in body["answer"]
    assert "approved rule row for site_cover" in " ".join(body["missing_information"])


def test_project_chat_refuses_targeted_check_question_without_latest_run(client):
    project = create_project(client)

    answer = client.post(
        f"/v1/projects/{project['id']}/chat",
        json={"message": "What did the site cover check say?"},
    )

    assert answer.status_code == 200, answer.text
    body = answer.json()
    assert body["status"] == "missing_info"
    assert body["citations"] == []
    assert "No completed deterministic compliance run exists" in body["answer"]
    assert "No completed compliance run exists for this project." in body["missing_information"]


def test_project_chat_refuses_build_question_without_latest_run_even_when_source_library_matches(client):
    seed_building_height_source(client)
    project = create_project(client)

    answer = client.post(
        f"/v1/projects/{project['id']}/chat",
        json={"message": "Can I build a two storey house in R40 in Cockburn?"},
    )

    assert answer.status_code == 200, answer.text
    body = answer.json()
    assert body["status"] == "missing_info"
    assert body["citations"] == []
    assert "No completed deterministic compliance run exists" in body["answer"]
    assert "single house may be two storeys" not in body["answer"]


def test_project_chat_summarises_latest_run_for_build_question_instead_of_source_library(client):
    seed_building_height_source(client)
    project = create_project(client)
    add_measurements(client, project["id"])

    run = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert run.status_code == 200, run.text

    answer = client.post(
        f"/v1/projects/{project['id']}/chat",
        json={"message": "Can I build a two storey house in R40 in Cockburn?"},
    )

    assert answer.status_code == 200, answer.text
    body = answer.json()
    assert body["status"] in {"missing_info", "unsupported"}
    assert body["citations"] == []
    assert "latest deterministic compliance run" in body["answer"]
    assert "single house may be two storeys" not in body["answer"]


def test_not_applicable_trigger_without_citation_is_unsupported(client):
    project = create_project(client)
    response = client.post(
        f"/v1/projects/{project['id']}/measurements",
        json={"key": "ancillary_dwelling_proposed", "value": 0, "unit": "count"},
    )
    assert response.status_code == 200, response.text

    run = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert run.status_code == 200, run.text
    ancillary = next(row for row in run.json()["results"] if row["check_key"] == "ancillary_dwelling_trigger")

    assert ancillary["status"] == "unsupported"
    assert "approved source citation" in ancillary["missing_information"]


def test_rfi_parse_refuses_weak_source_requirement_candidates(client):
    source = client.post(
        "/v1/sources/seed",
        json={
            "title": "Accepted Weak RFI Site Cover Note",
            "jurisdiction": "WA",
            "authority": "City of Cockburn",
            "local_government": "Cockburn",
            "source_type": "local_planning_policy",
            "canonical_url": "https://example.test/weak-rfi-site-cover-note",
            "licence_notes": "Public council policy fixture.",
            "access_type": "public",
            "review_status": "accepted",
            "content": "5.3.1 Open space\nOpen space and site cover calculations should be shown on the site plan.",
            "version_label": "accepted",
        },
    )
    assert source.status_code == 200, source.text
    project = create_project(client)

    rfi = client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={"text": "1. Please confirm the site cover requirement for R30."},
    )

    assert rfi.status_code == 200, rfi.text
    assert rfi.json()[0]["source_requirement_candidates"] == []

    draft = client.post(f"/v1/projects/{project['id']}/rfi/draft-response")
    assert draft.status_code == 200, draft.text
    item = draft.json()["content"]["item_table"][0]
    assert item["source_support_status"] == "unsupported"
    assert item["source_citation_count"] == 0


def test_rfi_response_export_and_audit_traceability(client):
    seed_source(client)
    project = create_project(client)
    rfi = client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={
            "text": "\n".join(
                [
                    "1. Could the garage be reoriented to reduce dominance and improve passive surveillance on A05?",
                    "2. Please confirm the front setback dimension and update the site plan.",
                ]
            )
        },
    )
    assert rfi.status_code == 200, rfi.text
    assert len(rfi.json()) == 2
    assert rfi.json()[0]["source_requirement_candidates"] == []

    draft = client.post(f"/v1/projects/{project['id']}/rfi/draft-response")
    assert draft.status_code == 200, draft.text
    draft_body = draft.json()
    assert "does not assert final compliance" in draft_body["draft_text"]
    assert "No approved source citation matched RFI item 1" in draft_body["draft_text"]
    assert draft_body["requires_human_review"] is True
    assert draft_body["liability_notice"].startswith("Draft only")
    assert "approved source citation for RFI item 1" in draft_body["missing_information"]
    assert draft_body["content"]["item_table"][0]["source_support_status"] == "unsupported"
    assert draft_body["content"]["item_table"][0]["source_citation_count"] == 0
    old_draft_id = draft_body["id"]

    later_rfi = client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={"text": "3. Please provide an updated landscaping note for the street setback area."},
    )
    assert later_rfi.status_code == 200, later_rfi.text
    assert later_rfi.json()[0]["item_number"] == 3
    latest_draft = client.post(f"/v1/projects/{project['id']}/rfi/draft-response")
    assert latest_draft.status_code == 200, latest_draft.text
    latest_draft_body = latest_draft.json()
    assert latest_draft_body["id"] != old_draft_id

    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["rfi_items", "rfi_response", "source_list"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    assert export.json()["file_sha256"]
    assert export.json()["status"] == "pending_human_signoff"
    assert export.json()["manifest"]["requires_human_signoff"] is True
    assert export.json()["manifest"]["submission_ready"] is False
    assert export.json()["manifest"]["human_signoff_status"] == "required"
    manifest = client.get(f"/v1/projects/{project['id']}/exports/{export.json()['id']}")
    assert manifest.status_code == 200
    assert manifest.json()["id"] == export.json()["id"]
    download = client.get(f"/v1/projects/{project['id']}/exports/{export.json()['id']}/download")
    assert download.status_code == 200
    assert download.headers["x-draftcheck-export-id"] == export.json()["id"]
    assert download.headers["x-draftcheck-human-signoff-status"] == "required"
    assert download.headers["x-draftcheck-submission-ready"] == "false"
    assert download.headers["content-type"].startswith("application/json")
    downloaded = download.json()
    assert downloaded["requires_human_signoff"] is True
    assert downloaded["source_signoff_notice"].lower().startswith("human signoff required")
    assert downloaded["sections"]["source_list"] == []
    assert [draft["id"] for draft in downloaded["sections"]["rfi_response"]] == [latest_draft_body["id"]]
    assert old_draft_id not in {draft["id"] for draft in downloaded["sections"]["rfi_response"]}
    assert "Item 3" in downloaded["sections"]["rfi_response"][0]["draft_text"]
    assert "updated landscaping note" in downloaded["sections"]["rfi_response"][0]["draft_text"]

    audit = client.get(f"/v1/audit?project_id={project['id']}")
    assert audit.status_code == 200
    actions = {event["action"] for event in audit.json()}
    assert {"project.created", "rfi.parsed", "response_draft.generated", "export.created"}.issubset(actions)

    jobs = [event for event in audit.json() if event["action"] == "job.enqueued"]
    assert jobs, "Hermes adapter should create traceable disabled jobs locally"


def test_docx_export_download_returns_word_file(client):
    project = create_project(client)
    client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={"text": "1. Please confirm the front setback dimension."},
    )
    client.post(f"/v1/projects/{project['id']}/rfi/draft-response")
    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "docx", "sections": ["rfi_response"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    download = client.get(f"/v1/projects/{project['id']}/exports/{export.json()['id']}/download")
    assert download.status_code == 200
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    document = Document(BytesIO(download.content))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "DraftCheck WA Response Pack" in text
    assert "Human Signoff" in text


def test_enabled_export_formats_include_human_signoff_notice(client):
    project = create_project(client)
    client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={"text": "1. Please confirm the front setback dimension."},
    )
    client.post(f"/v1/projects/{project['id']}/rfi/draft-response")

    for export_format in ["csv", "xlsx", "html"]:
        export = client.post(
            f"/v1/projects/{project['id']}/exports",
            json={"format": export_format, "sections": ["rfi_response"], "created_by": "tester"},
        )
        assert export.status_code == 200, export.text
        download = client.get(f"/v1/projects/{project['id']}/exports/{export.json()['id']}/download")
        assert download.status_code == 200

        if export_format == "xlsx":
            workbook = load_workbook(BytesIO(download.content))
            sheet_names = workbook.sheetnames
            assert "Human Signoff" in sheet_names
            sheet_text = "\n".join(str(cell.value or "") for row in workbook["Human Signoff"] for cell in row)
        else:
            sheet_text = download.content.decode("utf-8")

        assert "Human signoff required before submission" in sheet_text


def test_export_signoff_updates_manifest_and_download_gate(client):
    project = create_project(client)
    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["rfi_response"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    export_id = export.json()["id"]

    signoff = client.post(
        f"/v1/projects/{project['id']}/signoffs",
        json={
            "target_type": "export",
            "target_id": export_id,
            "status": "approved_for_export",
            "signed_by": "reviewer@example.test",
            "notes": "Reviewed for export packaging.",
        },
    )
    assert signoff.status_code == 200, signoff.text

    manifest = client.get(f"/v1/projects/{project['id']}/exports/{export_id}")
    assert manifest.status_code == 200, manifest.text
    manifest_body = manifest.json()
    assert manifest_body["status"] == "signed_off"
    assert manifest_body["manifest"]["requires_human_signoff"] is False
    assert manifest_body["manifest"]["submission_ready"] is True
    assert manifest_body["manifest"]["human_signoff_status"] == "approved_for_export"
    assert manifest_body["manifest"]["human_signoff_id"] == signoff.json()["id"]
    assert manifest_body["manifest"]["signed_by"] == "reviewer@example.test"
    assert "does not certify" in manifest_body["manifest"]["source_signoff_notice"]

    download = client.get(f"/v1/projects/{project['id']}/exports/{export_id}/download")
    assert download.status_code == 200
    assert download.headers["x-draftcheck-human-signoff-status"] == "approved_for_export"
    assert download.headers["x-draftcheck-submission-ready"] == "true"


def test_export_signoff_fails_closed_when_legacy_manifest_lacks_readiness_summary():
    export = Export(
        id="exp_legacy",
        project_id="proj_legacy",
        export_type="response_pack",
        format="json",
        status="pending_human_signoff",
        manifest_json=to_json(
            {
                "requires_human_signoff": True,
                "submission_ready": False,
                "human_signoff_status": "required",
            }
        ),
    )
    signoff = HumanSignoff(
        id="sig_legacy",
        project_id="proj_legacy",
        target_type="export",
        target_id="exp_legacy",
        status="approved_for_export",
        signed_by="reviewer@example.test",
        notes="Legacy export reviewed.",
        created_at=utcnow(),
    )

    _apply_export_signoff(export, signoff)

    manifest = from_json(export.manifest_json, {})
    readiness = manifest["export_readiness"]
    assert export.status == "signed_off_with_blockers"
    assert manifest["requires_human_signoff"] is True
    assert manifest["submission_ready"] is False
    assert manifest["human_signoff_status"] == "approved_for_export_with_blocking_issues"
    assert readiness["status"] == "blocked"
    assert readiness["blocking_issue_count"] == 1
    assert readiness["blocking_issues"][0]["section"] == "export_manifest"
    assert "regenerate the export" in readiness["blocking_issues"][0]["reason"]


def test_export_signoff_cannot_make_unresolved_compliance_submission_ready(client):
    project = create_project(client)
    add_measurements(client, project["id"])
    matrix = client.post(f"/v1/projects/{project['id']}/checks/run")
    assert matrix.status_code == 200, matrix.text
    assert any(row["status"] in {"missing_info", "needs_human_review", "unsupported"} for row in matrix.json()["results"])

    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["compliance_matrix"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    export_id = export.json()["id"]
    assert export.json()["manifest"]["export_readiness"]["status"] == "blocked"
    assert export.json()["manifest"]["export_readiness"]["blocking_issue_count"] > 0

    signoff = client.post(
        f"/v1/projects/{project['id']}/signoffs",
        json={
            "target_type": "export",
            "target_id": export_id,
            "status": "approved_for_export",
            "signed_by": "reviewer@example.test",
            "notes": "Reviewed unresolved compliance output.",
        },
    )
    assert signoff.status_code == 200, signoff.text

    manifest = client.get(f"/v1/projects/{project['id']}/exports/{export_id}")
    assert manifest.status_code == 200, manifest.text
    manifest_body = manifest.json()
    assert manifest_body["status"] == "signed_off_with_blockers"
    assert manifest_body["manifest"]["requires_human_signoff"] is True
    assert manifest_body["manifest"]["submission_ready"] is False
    assert manifest_body["manifest"]["human_signoff_status"] == "approved_for_export_with_blocking_issues"
    assert "not submission-ready" in manifest_body["manifest"]["source_signoff_notice"]

    download = client.get(f"/v1/projects/{project['id']}/exports/{export_id}/download")
    assert download.status_code == 200
    assert download.headers["x-draftcheck-human-signoff-status"] == "approved_for_export_with_blocking_issues"
    assert download.headers["x-draftcheck-submission-ready"] == "false"


def test_export_signoff_cannot_make_unresolved_rfi_items_submission_ready(client):
    project = create_project(client)
    parsed = client.post(
        f"/v1/projects/{project['id']}/rfi/parse",
        json={"text": "1. Please confirm the front setback dimension on the site plan."},
    )
    assert parsed.status_code == 200, parsed.text
    assert parsed.json()[0]["missing_evidence"]

    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["rfi_items"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    export_id = export.json()["id"]
    readiness = export.json()["manifest"]["export_readiness"]
    assert readiness["status"] == "blocked"
    assert any(issue["section"] == "rfi_items" for issue in readiness["blocking_issues"])

    signoff = client.post(
        f"/v1/projects/{project['id']}/signoffs",
        json={
            "target_type": "export",
            "target_id": export_id,
            "status": "approved_for_export",
            "signed_by": "reviewer@example.test",
            "notes": "Reviewed unresolved RFI item list.",
        },
    )
    assert signoff.status_code == 200, signoff.text

    manifest = client.get(f"/v1/projects/{project['id']}/exports/{export_id}")
    assert manifest.status_code == 200, manifest.text
    manifest_body = manifest.json()
    assert manifest_body["status"] == "signed_off_with_blockers"
    assert manifest_body["manifest"]["requires_human_signoff"] is True
    assert manifest_body["manifest"]["submission_ready"] is False
    assert manifest_body["manifest"]["human_signoff_status"] == "approved_for_export_with_blocking_issues"


def test_export_revision_signoff_keeps_submission_ready_false(client):
    project = create_project(client)
    export = client.post(
        f"/v1/projects/{project['id']}/exports",
        json={"format": "json", "sections": ["rfi_response"], "created_by": "tester"},
    )
    assert export.status_code == 200, export.text
    export_id = export.json()["id"]

    signoff = client.post(
        f"/v1/projects/{project['id']}/signoffs",
        json={
            "target_type": "export",
            "target_id": export_id,
            "status": "needs_revision",
            "signed_by": "reviewer@example.test",
            "notes": "Update source list before export.",
        },
    )
    assert signoff.status_code == 200, signoff.text

    manifest = client.get(f"/v1/projects/{project['id']}/exports/{export_id}")
    assert manifest.status_code == 200, manifest.text
    manifest_body = manifest.json()
    assert manifest_body["status"] == "needs_revision"
    assert manifest_body["manifest"]["requires_human_signoff"] is True
    assert manifest_body["manifest"]["submission_ready"] is False
    assert manifest_body["manifest"]["human_signoff_status"] == "needs_revision"


class RecordingStorage:
    def __init__(self, bucket: str):
        self.bucket = bucket
        self.objects: dict[str, bytes] = {}

    def put_bytes(self, key: str, content: bytes) -> StoredObject:
        object_key = f"s3://{self.bucket}/{key}"
        self.objects[object_key] = content
        return StoredObject(object_key=object_key, content_sha256="test-sha", byte_size=len(content))

    def get_bytes(self, key: str) -> bytes:
        return self.objects[key]

    def exists(self, key: str) -> bool:
        return key in self.objects
